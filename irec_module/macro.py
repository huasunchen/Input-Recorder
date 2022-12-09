
"""
Input Recorder - Record and play back keyboard and mouse input.
Copyright (C) 2022  Zuzu_Typ

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

try:
    from . import config, util
except ImportError:
    import config

import winput, time, ctypes, zlib, json
import pyautogui
import numpy as np
import base64, io
from PIL import Image
import cv2
from irec_module.util import is_similar_image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ODII

GLOBAL_W = 64
GLOBAL_H = 64
GLOBAL_HALF_W = int(GLOBAL_W/2)
GLOBAL_HALF_H = int(GLOBAL_H/2)
GLOBAL_STEP_ID = 0

float_to_bytes = lambda f: bytes(ctypes.c_float(f))
bytes_to_float = lambda b: ctypes.cast(b, ctypes.POINTER(ctypes.c_float)).contents.value

perf_counter_ns = None

if hasattr(time, "perf_counter_ns"):
    perf_counter_ns = time.perf_counter_ns
elif hasattr(time, "perf_counter"):
    perf_counter_ns = lambda: int(time.perf_counter() * 1000000000.)
else:
    perf_counter_ns = lambda: int(time.time() * 1000000000.)

class MacroConfig:
    version = 1
    
    def __init__(self, screen_width, screen_height):
        assert type(screen_width) == type(screen_height) == int
        
        self.screen_width = screen_width
        self.screen_height = screen_height

    def to_dict(self):
        return {
            "version"       : self.version,
            "screen_width"  : self.screen_width,
            "screen_height" : self.screen_height
        }

    def to_json(self):
        return json.dumps()

    def to_bytes(self):
        data = self.version.to_bytes(1, "little") + self.screen_width.to_bytes(2, "little") + self.screen_height.to_bytes(2, "little")

        return len(data).to_bytes(1, "little") + data

    @classmethod
    def from_dict(cls, as_dict):
        assert "version" in as_dict and \
               "screen_width" in as_dict and \
               "screen_height" in as_dict

        assert as_dict["version"] == cls.version

        assert type(as_dict["screen_width"]) == int and as_dict["screen_width"] > 0
        assert type(as_dict["screen_height"]) == int and as_dict["screen_height"] > 0

        return cls(as_dict["screen_width"], as_dict["screen_height"])

    @classmethod
    def from_json(cls, string):
        return cls.from_dict(json.loads(string))


    @classmethod
    def from_bytes(cls, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) > 1 and int.from_bytes(bytes_obj[0:1], "little") + 1 == len(bytes_obj)

        assert int.from_bytes(bytes_obj[1:2], "little") == cls.version

        return cls(int.from_bytes(bytes_obj[2:4], "little"), int.from_bytes(bytes_obj[4:], "little"))

class MacroEvent:
    def prepare(self):
        raise NotImplementedError()
    
    def execute(self):
        raise NotImplementedError()

    def to_bytes(self):
        raise NotImplementedError()

    def to_dict(self):
        raise NotImplementedError()

    def to_json(self):
        return json.dumps(self.to_dict())

    @classmethod
    def from_bytes(cls):
        raise NotImplementedError()

    @classmethod
    def from_dict(cls):
        raise NotImplementedError()

    @classmethod
    def from_json(cls, config, string):
        return cls.from_dict(config, json.loads(string))

class MousePositionEvent(MacroEvent):
    bytecode = b"P"
    
    def __init__(self, config, x, y):
        self.config = config
        self.rel_x = float(x) / config.screen_width
        self.rel_y = float(y) / config.screen_height

    def copy(self):
        self.prepare()
        return self.__class__(self.config, self.x, self.y)

    def prepare(self):
        self.x = int(round(self.config.screen_width * self.rel_x, 0))
        self.y = int(round(self.config.screen_height * self.rel_y, 0))

    def execute(self):
        winput.set_mouse_pos(self.x, self.y)

    def to_dict(self):
        self.prepare()
        return {
            "type"  : self.__class__.__name__,
            "x"     : self.x,
            "y"     : self.y
        }

    def to_bytes(self):
        self.prepare()
        return self.bytecode + self.x.to_bytes(2, "little", signed=True) + self.y.to_bytes(2, "little", signed=True)

    @classmethod
    def from_bytes(cls, config, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) == 5 and bytes_obj[0:1] == cls.bytecode

        return cls(config, int.from_bytes(bytes_obj[1:3], "little", signed=True), int.from_bytes(bytes_obj[3:], "little", signed=True))

    @classmethod
    def from_dict(cls, config, as_dict):
        assert "type" in as_dict and \
               "x" in as_dict and \
               "y" in as_dict

        assert as_dict["type"] == cls.__name__

        assert type(as_dict["x"]) == int
        assert type(as_dict["y"]) == int

        return cls(config, as_dict["x"], as_dict["y"])

    def __str__(self):
        return "Set mouse position to ({:.2f}%, {:.2f}%)".format(self.rel_x * 100, self.rel_y * 100)

### doesn't work right because of the mouse acceleration
##class MouseMoveEvent(MacroEvent):
##    bytecode = b"M"
##    
##    def __init__(self, config, dx, dy):
##        self.config = config
##        self.rel_dx = float(dx) / config.screen_width
##        self.rel_dy = float(dy) / config.screen_height
##
##    def prepare(self):
##        self.dx = int(round(self.config.screen_width * self.rel_dx, 0))
##        self.dy = int(round(self.config.screen_height * self.rel_dy, 0))
##
##    def execute(self):
##        winput.move_mouse(self.dx, self.dy)
##
##    def to_bytes(self):
##        self.prepare()
##        return self.bytecode + self.dx.to_bytes(2, "little", signed=True) + self.dy.to_bytes(2, "little", signed=True)
##
##    @classmethod
##    def from_bytes(cls, config, bytes_obj):
##        assert type(bytes_obj) == bytes and len(bytes_obj) == 5 and bytes_obj[0:1] == cls.bytecode
##
##        return cls(config, int.from_bytes(bytes_obj[1:3], "little", signed=True), int.from_bytes(bytes_obj[3:], "little", signed=True))
##
##    def __str__(self):
##        return "Move mouse by ({:.2f}%, {:.2f}%)".format(self.rel_dx, self.rel_dy)

class MouseWheelEvent(MacroEvent):
    def __init__(self, amount):
        self.amount = amount

    def copy(self):
        return self.__class__(self.amount)

    def prepare(self):
        pass

    def to_dict(self):
        return {
            "type"  : self.__class__.__name__,
            "amount": self.amount
        }

    def to_bytes(self):
        return self.bytecode + self.amount.to_bytes(1, "little", signed=True)

    @classmethod
    def from_bytes(cls, config, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) == 2 and bytes_obj[0:1] == cls.bytecode

        return cls(int.from_bytes(bytes_obj[1:], "little", signed=True))

    @classmethod
    def from_dict(cls, config, as_dict):
        assert "type" in as_dict and \
               "amount" in as_dict

        assert as_dict["type"] == cls.__name__

        assert type(as_dict["amount"]) == int

        return cls(as_dict["amount"])

class MouseWheelMoveEvent(MouseWheelEvent):
    bytecode = b"V"
    
    def execute(self):
        winput.move_mousewheel(self.amount)

    def __str__(self):
        return "Move mousewheel by {:+1d}".format(int(self.amount))

class MouseWheelHorizontalMoveEvent(MouseWheelEvent):
    bytecode = b"H"
    
    def execute(self):
        winput.move_mousewheel(self.amount, True)

    def __str__(self):
        return "Move mousewheel horizontally by {:+1d}".format(int(self.amount))

class MouseButtonEvent(MacroEvent):
    def __init__(self, mouse_button):
        self.mouse_button = mouse_button

    def copy(self):
        return self.__class__(self.mouse_button)

    def prepare(self):
        pass

    def to_dict(self):
        return {
            "type"  : self.__class__.__name__,
            "mouse_button": self.mouse_button
        }

    def to_bytes(self):
        return self.bytecode + self.mouse_button.to_bytes(1, "little")

    @classmethod
    def from_bytes(cls, config, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) == 2 and bytes_obj[0:1] == cls.bytecode

        return cls(int.from_bytes(bytes_obj[1:], "little"))

    @classmethod
    def from_dict(cls, config, as_dict):
        assert "type" in as_dict and \
               "mouse_button" in as_dict

        assert as_dict["type"] == cls.__name__

        assert type(as_dict["mouse_button"]) == int

        return cls(as_dict["mouse_button"])

class MouseButtonPressEvent(MouseButtonEvent):
    bytecode = b"B"

    def __init__(self, mouse_button, region_data, x, y):
        super().__init__(mouse_button)
        self.region_data = region_data
        self.x = x
        self.y = y
        # print("__init__={}".format(self.region_data))

    def to_bytes(self):
        return self.bytecode + self.mouse_button.to_bytes(1, "little") + self.x.to_bytes(2, "little")  + self.y.to_bytes(2, "little") + self.region_data.tobytes()

    def to_dict(self):
        return {
            "type"  : self.__class__.__name__,
            "mouse_button": self.mouse_button,
            "x": self.x,
            "y": self.y,
            "img": base64.b64encode(self.region_data.tobytes()).decode()
        }

    @classmethod
    def from_dict(cls, config, as_dict):
        assert "type" in as_dict and \
               "mouse_button" in as_dict and \
               "img" in as_dict

        assert as_dict["type"] == cls.__name__
        assert type(as_dict["mouse_button"]) == int
        if as_dict["img"]:
            img_decoded = base64.b64decode(as_dict["img"])
            img_data = np.reshape(np.frombuffer(img_decoded, dtype=np.uint8), newshape=(GLOBAL_W, GLOBAL_H, 3))
        else:
            img_data = []
        return cls(as_dict["mouse_button"], img_data, as_dict["x"], as_dict["y"])

    @classmethod
    def from_bytes(cls, config, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) >= 2 and bytes_obj[0:1] == cls.bytecode
        deserialized_bytes = np.frombuffer(bytes_obj[6:], dtype=np.uint8)
        deserialized = np.reshape(deserialized_bytes, newshape=(GLOBAL_W, GLOBAL_H, 3))
        # print("deserialized={}".format(deserialized))
        return MouseButtonPressEvent(int.from_bytes(bytes_obj[1:2], "little"), deserialized, int.from_bytes(bytes_obj[2:4], "little"), int.from_bytes(bytes_obj[4:6], "little"))

    @staticmethod
    def is_match(rd, fd, sx, sy, w, h):
        return np.array_equal(rd, fd[sy:sy+h, sx:sx+w])


    @staticmethod
    def check_color(full_img, regn_img):
        # import cv2
        regn_data = cv2.cvtColor(regn_img, cv2.COLOR_BGR2RGB)
        full_data = cv2.cvtColor(full_img, cv2.COLOR_BGR2RGB)
        diff_num = 0;
        for x in range(0, GLOBAL_W):
            for y in range(0, GLOBAL_H):
                regn_pixel = regn_data[y, x]
                full_pixel = full_data[y, x]
                diff_r = int(full_pixel[0]) - int(regn_pixel[0])
                diff_g = int(full_pixel[1]) - int(regn_pixel[1])
                diff_b = int(full_pixel[2]) - int(regn_pixel[2])
                DIFF = 20
                if abs(diff_r) > DIFF or abs(diff_g) > DIFF or abs(diff_b) > DIFF:
                    ++diff_num
        return diff_num < (GLOBAL_W * GLOBAL_H) * 0.1

    @staticmethod
    def get_mouse_position(mouse_x0, mouse_y0):
        # import cv2

        sift = cv2.xfeatures2d.SIFT_create()
        full_img = cv2.imread('match_full.png', 0)
        regn_img = cv2.imread('match_regn.png', 0)

        kp_full, desc_full = sift.detectAndCompute(full_img, None)
        kp_regn, desc_regn = sift.detectAndCompute(regn_img, None)

        if len(kp_full) == 0 or len(kp_regn) == 0:
            if MouseButtonPressEvent.check_color(full_img, regn_img):
                return (mouse_x0, mouse_y0)
            return None

        bf = cv2.BFMatcher(crossCheck=True)
        matches = bf.match(desc_full, desc_regn)
        matches = sorted(matches, key=lambda x: x.distance)
        match_position = kp_full[matches[0].queryIdx].pt if len(matches) > 0 else (-1,-1)
        # match_distance = matches[0].distance
        # match_dist2 = pow(match_position[0] - mouse_x0,2) + pow(match_position[1] - mouse_y0,2)
        # for mm in matches:
        #     if abs(mm.distance - match_distance) > 3:
        #         continue
        #     pt = kp_full[mm.queryIdx].pt
        #     dist2 = pow(pt[0] - mouse_x0,2) + pow(pt[1] - mouse_y0,2)
        #     if dist2 < match_dist2:
        #         match_dist2 = dist2
        #         match_position = pt

        print("match distance=", matches[0].distance, "pt=", match_position, "mouse_t=", (mouse_x0, mouse_y0))

        if matches[0].distance < 150:
            return (mouse_x0, mouse_y0)
        return None

    def execute(self):
        import keyboard
        global continue_playback
        global GLOBAL_STEP_ID

        found = False
        GLOBAL_STEP_ID = GLOBAL_STEP_ID + 1
        # time.sleep(0.2)
        mouse_x0 = self.x
        mouse_y0 = self.y
        if len(self.region_data) > 0:
            regn_img = Image.fromarray(np.uint8(self.region_data))
            regn_img.save('match_regn.png')

            for zz in range(0,1000):
                if keyboard.is_pressed("esc"):
                    full_img = pyautogui.screenshot(region=[mouse_x0 - GLOBAL_HALF_W, mouse_y0 - GLOBAL_HALF_H, GLOBAL_W, GLOBAL_H])  # x,y,w,h
                    regn_img.save('match_{}_regn.png'.format(GLOBAL_STEP_ID))
                    full_img.save('match_{}_full.png'.format(GLOBAL_STEP_ID))
                    exit(1)
                    return
                full_img = pyautogui.screenshot(region=[mouse_x0 - GLOBAL_HALF_W, mouse_y0 - GLOBAL_HALF_H, GLOBAL_W, GLOBAL_H]) # x,y,w,h
                full_img.save("match_full.png")
                mouse_should_pos = MouseButtonPressEvent.get_mouse_position(mouse_x0, mouse_y0)
                if mouse_should_pos == None:
                    # regn_img2 = Image.fromarray(np.uint8(self.region_data))
                    # regn_img2.save('test_match/screenshot_regn.png')
                    # full_img.save('test_match/screenshot_full.png')
                    print("mouse position no matched! mouse pos={}".format((mouse_x0, mouse_y0)))
                    time.sleep(0.05)
                    continue
                winput.set_mouse_pos(self.x, self.y)
                winput.press_mouse_button(self.mouse_button)
                found = True
                break
            if not found:
                full_img = pyautogui.screenshot(region=[mouse_x0 - GLOBAL_HALF_W, mouse_y0 - GLOBAL_HALF_H, GLOBAL_W, GLOBAL_H])  # x,y,w,h
                regn_img.save('match_{}_regn.png'.format(GLOBAL_STEP_ID))
                full_img.save('match_{}_full.png'.format(GLOBAL_STEP_ID))
                winput.set_mouse_pos(self.x, self.y)
                winput.press_mouse_button(self.mouse_button)
        else:
            winput.set_mouse_pos(self.x, self.y)
            winput.press_mouse_button(self.mouse_button)

    def __str__(self):
        return "Press {} mouse button".format(util.mouse_button_to_name(self.mouse_button).lower())
        
class MouseButtonReleaseEvent(MouseButtonEvent):
    bytecode = b"b"
    
    def execute(self):
        winput.release_mouse_button(self.mouse_button)

    def __str__(self):
        return "Release {} mouse button".format(util.mouse_button_to_name(self.mouse_button).lower())

class KeyEvent(MacroEvent):
    def __init__(self, vk_code):
        self.vk_code = vk_code

    def copy(self):
        return self.__class__(self.vk_code)

    def prepare(self):
        pass

    def to_dict(self):
        return {
            "type"  : self.__class__.__name__,
            "vk_code": self.vk_code
        }

    def to_bytes(self):
        return self.bytecode + self.vk_code.to_bytes(2, "little")

    @classmethod
    def from_bytes(cls, config, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) == 3 and bytes_obj[0:1] == cls.bytecode

        return cls(int.from_bytes(bytes_obj[1:], "little"))

    @classmethod
    def from_dict(cls, config, as_dict):
        assert "type" in as_dict and \
               "vk_code" in as_dict

        assert as_dict["type"] == cls.__name__

        assert type(as_dict["vk_code"]) == int

        return cls(as_dict["vk_code"])

class KeyPressEvent(KeyEvent):
    bytecode = b"K"
    
    def execute(self):
        winput.press_key(self.vk_code)

    def __str__(self):
        return "Press {}".format(util.vk_code_to_key_name(self.vk_code))

class KeyReleaseEvent(KeyEvent):
    bytecode = b"k"
    
    def execute(self):
        winput.release_key(self.vk_code)

    def __str__(self):
        return "Release {}".format(util.vk_code_to_key_name(self.vk_code))

event_bytecode_class_map = {
        b"P" : MousePositionEvent,
##        b"M" : MouseMoveEvent,
        b"V" : MouseWheelMoveEvent,
        b"H" : MouseWheelHorizontalMoveEvent,
        b"B" : MouseButtonPressEvent,
        b"b" : MouseButtonReleaseEvent,
        b"K" : KeyPressEvent,
        b"k" : KeyReleaseEvent,
    }

class EventExecutor:
    def __init__(self, time_offset, event):
        assert type(time_offset) == int and isinstance(event, MacroEvent)
        
        self.time_offset = time_offset
        self.event = event

    def copy(self):
        return EventExecutor(self.time_offset, self.event.copy())

    def prepare(self):
        self.event.prepare()

    def execute_at_time_offset(self, start_time):
        global continue_playback, enable_playback_interruption
        now = perf_counter_ns()

        target_time = start_time + self.time_offset

        while now < target_time:
            diff_in_ms = (target_time - now) // 1000000

            if diff_in_ms > 10:
                time.sleep(diff_in_ms // 10 / 100.)

            else:
                time.sleep(1 / 1000.)

            if not continue_playback:
                return 0

            if enable_playback_interruption:
                winput.get_message()
                
            now = perf_counter_ns()

        if not continue_playback:
            return 0

        stime = perf_counter_ns()
        self.event.execute()
        dtime = perf_counter_ns() - stime
        return dtime

    def to_dict(self):
        return {
            "time_offset"   : self.time_offset,
            "event"         : self.event.to_dict()
        }

    def to_json(self):
        return json.dumps(self.to_dict())

    def to_bytes(self):
        data = self.time_offset.to_bytes(8, "little") + self.event.to_bytes()

        return len(data).to_bytes(2, "little") + data

    @classmethod
    def from_bytes(cls, config, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) >= 10 and int.from_bytes(bytes_obj[0:2], "little") + 2 == len(bytes_obj)
        
        time_offset = int.from_bytes(bytes_obj[2:10], "little")

        event_type = bytes_obj[10:11]

        assert event_type in event_bytecode_class_map, "Unknown event bytecode '{}'".format(event_type)

        event_class = event_bytecode_class_map[event_type]

        event = event_class.from_bytes(config, bytes_obj[10:])

        return cls(time_offset, event)

    @classmethod
    def from_dict(cls, config, as_dict):
        assert "time_offset" in as_dict and \
               "event" in as_dict

        time_offset = as_dict["time_offset"]
        event_dict = as_dict["event"]

        assert type(time_offset) == int and time_offset >= 0

        assert "type" in event_dict

        glob = globals()

        assert event_dict["type"] in glob, "Unknown event type: {}".format(event_dict["type"])

        event_class = glob[event_dict["type"]]
        event = event_class.from_dict(config, event_dict)

        return cls(time_offset, event)

    @classmethod
    def from_json(cls, config, string):
        return cls.from_dict(config, json.loads(string))

    def __str__(self):
        return "Executing {} at {}".format(str(self.event), self.time_offset)

class Macro:
    def __init__(self, name, config, event_executor_list):
        self.name = name
        self.config = config
        self.event_executor_list = event_executor_list

    def stop_playback_callback(self, key_event):
        global continue_playback
        
        if key_event.vkCode == self.stop_playback_vk_code:
            continue_playback = False

    def run(self):
        global continue_playback, enable_playback_interruption
        
        for executor in self.event_executor_list:
            executor.prepare()

        continue_playback = True

        enable_playback_interruption = config.get("enable_stop_playback_key", False)

        if enable_playback_interruption:
            self.stop_playback_vk_code = config.get("stop_playback_key", winput.VK_ESCAPE)
            winput.hook_keyboard(self.stop_playback_callback)
            
        start_time = perf_counter_ns()
        for executor in self.event_executor_list:
            use_time = executor.execute_at_time_offset(start_time)
            start_time += use_time
            # if use_time > 1000:
            #     print(use_time)

            if not continue_playback:
                break

        if enable_playback_interruption:
            winput.unhook_keyboard()

    def to_bytes(self):
        name_encoded = self.name.encode()
        name_data = len(name_encoded).to_bytes(1, "little") + name_encoded
        return name_data + self.config.to_bytes() + b"".join(map(lambda ee: ee.to_bytes(), self.event_executor_list))

    def as_compressed_bytes(self):
        return zlib.compress(self.to_bytes(), 9)
    
    def to_dict(self):
        return {
            "name"   : self.name,
            "config" : self.config.to_dict(),
            "event_executor_list" : list(map(lambda x: x.to_dict(), self.event_executor_list))
        }

    def to_json(self):
        return json.dumps(self.to_dict())

    @classmethod
    def from_dict(cls, as_dict):
        assert "name" in as_dict and \
               "config" in as_dict and \
               "event_executor_list" in as_dict

        name = as_dict["name"]

        config_dict = as_dict["config"]

        event_executor_dict_list = as_dict["event_executor_list"]

        assert type(name) == str and \
               type(config_dict) == dict and \
               type(event_executor_dict_list) == list

        config = MacroConfig.from_dict(config_dict)

        event_executor_list = list(map(lambda x: EventExecutor.from_dict(config, x), event_executor_dict_list))

        return cls(name, config, event_executor_list)

    @classmethod
    def from_json(cls, string):
        return cls.from_dict(json.loads(string))

    @classmethod
    def from_bytes(cls, bytes_obj):
        assert type(bytes_obj) == bytes and len(bytes_obj) > 1 and len(bytes_obj) > int.from_bytes(bytes_obj[0:1], "little")

        offset = bytes_obj[0] + 1

        name = bytes_obj[1:offset].decode()

        config_data_length = bytes_obj[offset] + 1

        config_data = bytes_obj[offset:offset+config_data_length]

        config = MacroConfig.from_bytes(config_data)

        offset += config_data_length

        event_executor_list = []

        while offset < len(bytes_obj):
            event_data_length = int.from_bytes(bytes_obj[offset:offset+2], "little") + 2

            assert offset + event_data_length <= len(bytes_obj)

            event_data = bytes_obj[offset : offset + event_data_length]

            event_executor_list.append(EventExecutor.from_bytes(config, event_data))

            offset += event_data_length

        return Macro(name, config, event_executor_list)

    @classmethod
    def from_compressed_bytes(cls, compressed_bytes_obj):
        return cls.from_bytes(zlib.decompress(compressed_bytes_obj))

    @staticmethod
    def from_raw_data(name, start_time, start_mouse_pos, screen_width, screen_height, raw_data):
        event_executor_list = []

        macro_config = MacroConfig(screen_width, screen_height)

        if start_mouse_pos:
            event_executor_list.append(EventExecutor(0, MousePositionEvent(macro_config, start_mouse_pos[0], start_mouse_pos[1])))

        last_mouse_pos = start_mouse_pos

        for timestamp, raw_event in raw_data:
            time_offset = timestamp - start_time

            event = None

            if type(raw_event) == winput.MouseEvent:
                if raw_event.action == winput.WM_MOUSEMOVE:
                    event = MousePositionEvent(macro_config, raw_event.position[0], raw_event.position[1])
                    last_mouse_pos = raw_event.position

                elif raw_event.action == winput.WM_LBUTTONDOWN:
                    event = MouseButtonPressEvent(winput.LEFT_MOUSE_BUTTON, raw_event.additional_data, last_mouse_pos[0], last_mouse_pos[1])
 

                elif raw_event.action == winput.WM_LBUTTONUP:
                    event = MouseButtonReleaseEvent(winput.LEFT_MOUSE_BUTTON)

                elif raw_event.action == winput.WM_RBUTTONDOWN:
                    event = MouseButtonPressEvent(winput.RIGHT_MOUSE_BUTTON, np.uint8([]), last_mouse_pos[0], last_mouse_pos[1])

                elif raw_event.action == winput.WM_RBUTTONUP:
                    event = MouseButtonReleaseEvent(winput.RIGHT_MOUSE_BUTTON)

                elif raw_event.action == winput.WM_MBUTTONDOWN:
                    event = MouseButtonPressEvent(winput.MIDDLE_MOUSE_BUTTON, np.uint8([]), last_mouse_pos[0], last_mouse_pos[1])

                elif raw_event.action == winput.WM_MBUTTONUP:
                    event = MouseButtonReleaseEvent(winput.MIDDLE_MOUSE_BUTTON)

                elif raw_event.action == winput.WM_XBUTTONDOWN:
                    event = MouseButtonPressEvent(getattr(winput, "XMB" + str(raw_event.additional_data)), np.uint8([]), last_mouse_pos[0], last_mouse_pos[1])

                elif raw_event.action == winput.WM_XBUTTONUP:
                    event = MouseButtonReleaseEvent(getattr(winput, "XMB" + str(raw_event.additional_data)))

                elif raw_event.action == winput.WM_MOUSEWHEEL:
                    event = MouseWheelMoveEvent(int(raw_event.additional_data))

                elif raw_event.action == winput.WM_MOUSEHWHEEL:
                    event = MouseWheelHorizontalMoveEvent(int(raw_event.additional_data))

                else:
                    raise Exception("Unknown Mouse Event")

            elif type(raw_event) == winput.KeyboardEvent:
                if raw_event.action == winput.WM_KEYDOWN or raw_event.action == winput.WM_SYSKEYDOWN:
                    event = KeyPressEvent(raw_event.vkCode)
                    
                elif raw_event.action == winput.WM_KEYUP or raw_event.action == winput.WM_SYSKEYUP:
                    event = KeyReleaseEvent(raw_event.vkCode)

                else:
                    raise Exception("Unknown Keyboard Event")

            else:
                raise Exception("Unknown Event Type")

            event_executor_list.append(EventExecutor(time_offset, event))

        return Macro(name, macro_config, event_executor_list)

def callback(event):
    global raw_data

    if type(event) == winput.MouseEvent:
        if event.action == winput.WM_LBUTTONDOWN:
            beg_x = max(0,event.position[0] - GLOBAL_HALF_W)
            beg_y = max(0,event.position[1] - GLOBAL_HALF_H)
            img = pyautogui.screenshot(region=[beg_x, beg_y, GLOBAL_W, GLOBAL_H])
            event.additional_data = np.asarray(img)
    raw_data.append((perf_counter_ns(), event))

def callback_with_stop_key(event):
    global stop_recording_key

    if event.vkCode == stop_recording_key:
        winput.stop()

    else:
        global raw_data

        raw_data.append((perf_counter_ns(), event))

def callback_only_stop_key(event):
    global stop_recording_key

    if event.vkCode == stop_recording_key:
        winput.stop()


def create_macro(name, start_at, screen_width, screen_height):
    global start, raw_data, stop_recording_key

    mode = config.get("recording_mode", "key")

    duration = config.get("recording_duration", 10)

    stop_recording_key = config.get("recording_stop_key", winput.VK_ESCAPE)

    mouse_enabled = config.get("record_mouse", True)

    keyboard_enabled = config.get("record_keyboard", True)

    assert mode in ("timer", "key")

    assert type(duration) in (float, int) and duration > 0

    assert type(stop_recording_key) == int and 0 <= stop_recording_key <= 2**15

    assert type(mouse_enabled) == type(keyboard_enabled) == bool

    assert mouse_enabled or keyboard_enabled

    raw_data = []

    while True:
        if time.time() > start_at:
            break
        time.sleep(0.0001)

    start = perf_counter_ns()

    start_mouse_pos = None

    if mode == "timer":
        if mouse_enabled:
            start_mouse_pos = winput.get_mouse_pos()
            winput.hook_mouse(callback)

        if keyboard_enabled:
            winput.hook_keyboard(callback)
        
        while True:
            now = perf_counter_ns()

            if now >= start + duration * 1000000000:
                break

            winput.get_message()

    elif mode == "key":
        if mouse_enabled:
            start_mouse_pos = winput.get_mouse_pos()
            winput.hook_mouse(callback)

        if keyboard_enabled:
            winput.hook_keyboard(callback_with_stop_key)
        else:
            winput.hook_keyboard(callback_only_stop_key)
        
        winput.wait_messages()

    winput.unhook_mouse()
    winput.unhook_keyboard()

    return Macro.from_raw_data(name, start, start_mouse_pos, screen_width, screen_height, raw_data)

def macros_to_json(*macros, indent=None):
    assert macros and all(map(lambda m: isinstance(m, Macro), macros))

    macros_dict_list = list(map(lambda mcr: mcr.to_dict(), macros))

    return json.dumps(macros_dict_list, indent=indent)

def macros_from_json(string):
    assert type(string) == str

    macros_dict_list = json.loads(string)

    return list(map(lambda md: Macro.from_dict(md), macros_dict_list))

def macros_to_excel(file_name, *macros):
    assert macros and all(map(lambda m: isinstance(m, Macro), macros))
    macros_dict = list(map(lambda mcr: mcr.to_dict(), macros))[0]
    title_name = macros_dict["name"]
    event_name = "event_executor_list"
    wb = Workbook()
    ws = wb.active
    ws.title = title_name
    for key, value in macros_dict["config"].items():
        ws.append([key, value])
    ws1 = wb.create_sheet(title=event_name, index=1)
    colu_list = []
    rows_list = []
    pic_list = []
    for event in macros_dict[event_name]:
        tmp = []
        for k, v in event.items():
            if k != "event" and k not in colu_list:
                colu_list.append(k)
                colu_list.append("verifyPicture")
            elif k == "event":
                for x, y in v.items():
                    if x not in colu_list:
                        colu_list.append(x)
        for key, value in event.items():
            if key != "event":
                tmp.append(value)
                tmp.append(True)
                continue
            tmp.append(value["type"])
            tmp.append(value["x"] if "x" in value else None)
            tmp.append(value["y"] if "y" in value else None)
            tmp.append(value["mouse_button"] if "mouse_button" in value else None)
            if "img" in value:
                base64_data_str = value["img"]
                base64_data_bytes = base64_data_str.encode('utf-8')
                img_decoded = base64.b64decode(base64_data_bytes)
                nparr = np.reshape(np.frombuffer(img_decoded, dtype=np.uint8), newshape=(GLOBAL_W, GLOBAL_H, 3))
                segment_data = cv2.cvtColor(nparr, cv2.COLOR_BGR2RGB)  # IMREAD_GRAYSCALE
                regn_img = Image.fromarray(np.uint8(segment_data))
                regn_img.save(f'.tmp\\{value["x"]}.png')
                pic_list.append(f'.tmp\\{value["x"]}.png')
            else:
                pic_list.append(None)
        rows_list.append(tmp)
    ws1.append(colu_list)
    for x in rows_list:
        ws1.append(x)
    wb.save(file_name)
    wb1 = load_workbook(file_name)
    sheet = wb1[event_name]
    star_num = 2
    for pic in pic_list:
        if pic:
            img = ODII(pic)
            sheet.add_image(img, f'G{star_num}')
        else:
            sheet[f'G{star_num}'] = pic
        star_num += 1
    wb1.save(file_name)

def macros_from_excel(file_name):
    xlsx = load_workbook(file_name)
    sheet_names = xlsx.sheetnames
    macros_data = {}
    for sheet_name in sheet_names:
        ws = xlsx[sheet_name]
        if "_" not in sheet_name:
            macros_data["name"] = sheet_name
            col_data = list(ws.iter_cols(values_only=True))
            col_key = col_data[0]
            col_value = col_data[1:]
            for col in col_value:
                macros_data["config"] = dict(zip(col_key, col))
        else:
            cases_list = []
            datas = list(ws.iter_rows(values_only=True))
            case_title = datas[0]
            # case_title.reomove("verifyPicture")
            case_datas = datas[1:]
            for case in case_datas:
                result = dict(zip(case_title, case))
                tmp_dict = {}
                for k, v in result.items():
                    if k == "time_offset":
                        tmp_dict[k] = v
                    elif k == "verifyPicture" and v:
                        continue
                    elif k == "verifyPicture" and not v:
                        tmp_dict.setdefault("event", {})["img"] = v
                    elif k == "mouse_button" and v == None:
                        continue
                    elif k == "img" and v == None:
                        continue
                    elif k == "x" and v == None:
                        continue
                    elif k == "y" and v == None:
                        continue
                    else:
                        tmp_dict.setdefault("event", {})[k] = v
                cases_list.append(tmp_dict)
            _images = {}
            for image in ws._images:
                row = image.anchor._from.row
                img = Image.open(io.BytesIO(image._data()))
                _images[row - 1] = base64.b64encode(np.array(img).tobytes()).decode()
            for index, value in _images.items():
                if "img" in cases_list[index]["event"]:
                    continue
                cases_list[index]["event"]["img"] = value
            macros_data[sheet_name] = cases_list
    xlsx.close()
    macros_dict_list = json.loads(json.dumps([macros_data], indent=4))
    return list(map(lambda md: Macro.from_dict(md), macros_dict_list))

def macros_to_bytes(*macros, compressionlevel=9):
    assert macros and all(map(lambda m: isinstance(m, Macro), macros))

    data_array = []

    if len(macros) == 1:
        data_array.append(b"S")
        data_array.append(macros[0].to_bytes())

    else:
        data_array.append(b"M")
        data_array.append(len(macros).to_bytes(1, "little"))

        for macro in macros:
            macro_data = macro.to_bytes()
            data_array.append(len(macro_data).to_bytes(4, "little"))
            data_array.append(macro_data)

    uncompressed_data = b"".join(data_array)

    out_data = zlib.compress(uncompressed_data, compressionlevel)

    uncompressed_data_len = len(uncompressed_data)

    return uncompressed_data_len.to_bytes(4, "little") + out_data

def macros_from_bytes(bytes_obj):
    assert type(bytes_obj) == bytes

    content = bytes_obj[4:]

    uncompressed_bytes = zlib.decompress(content)

    assert uncompressed_bytes and uncompressed_bytes[0:1] in (b"S", b"M")

    if uncompressed_bytes[0:1] == b"S":
        return [Macro.from_bytes(uncompressed_bytes[1:])]

    else:
        macros = []
        
        offset = 1

        macro_count = uncompressed_bytes[offset]

        offset += 1

        while offset < len(uncompressed_bytes):
            assert len(uncompressed_bytes) > offset + 3

            macro_data_length = int.from_bytes(uncompressed_bytes[offset : offset + 4], "little")
            offset += 4
            macro_data = uncompressed_bytes[offset : offset + macro_data_length]
            
            macros.append(Macro.from_bytes(macro_data))

            offset += macro_data_length

        assert len(macros) == macro_count

        return macros

def request_key_callback(event):
    global requested_key
    requested_key = event.vkCode

def request_key(timeout):
    global requested_key
    
    requested_key = None
    
    start = time.time()

    winput.hook_keyboard(request_key_callback)

    while not requested_key:
        now = time.time()

        if now >= start + timeout:
            break

        winput.get_message()

    winput.unhook_keyboard()

    return requested_key

def request_mouse_callback(event):
    global last_request_position, last_request_timestamp

    last_request_position = event.position

    last_request_timestamp = time.time()

def request_mouse_pos(timeout):
    global last_request_position, last_request_timestamp

    last_request_position = None

    last_request_timestamp = time.time()

    winput.hook_mouse(request_mouse_callback)

    while True:
        now = time.time()

        winput.get_message()

        if last_request_position and last_request_timestamp + timeout < now:
            break

    winput.unhook_mouse()

    return last_request_position

    




























